import tkinter as tk
from tkinter import ttk
import openpyxl as xl
from datetime import datetime
import os
from tkinter import Canvas
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import glob
import statsmodels.api as sm

# Create the main window
root = tk.Tk()
root.title("The System")
root.geometry("1800x1600")
root.configure(bg="light cyan")
root.grid_rowconfigure(0 ) # this needed to be added
root.grid_columnconfigure(0) # as did this




#Load the Excel File
if os.path.exists('active.xlsx'):
    wb = xl.load_workbook("active.xlsx")
    per_sheet = wb['Personality']
#Create Excel File if it doesn't exist.
else: 
    wb=xl.Workbook()
    per_sheet = wb.create_sheet('Personality')
    sheet = wb["Sheet"]
    wb.remove(sheet) #Remove pre-existing Sheet

    # Add the Personality Traits to the first row
    per_sheet['A1'] = 'Date'
    per_sheet['B1'] = 'Smart'
    per_sheet['C1'] = 'Disciplined'
    per_sheet['D1'] = 'Meticulous'
    per_sheet['E1'] = 'Reserved'
    per_sheet['F1'] = 'Calm'
    per_sheet['G1'] = 'Kind'
    per_sheet['H1'] = 'Average'
    per_sheet['I1'] = 'Pandexity'
    per_sheet['J1'] = 'Happiness'
    wb.save("active.xlsx")


x = datetime.now()
current_year = x.year
current_date = x.strftime("%d %B")

if os.path.exists(f'{current_year}.xlsx'):
    wb_year = xl.load_workbook(f"{current_year}.xlsx")
#Create Excel File if it doesn't exist.
else: 
    wb_year=xl.Workbook()
    wb_year.save(f'{current_year}.xlsx')

def run():
    refresh_window = tk.Frame(root)
    refresh_window.grid(row=0,column=0)
    upper_section = tk.Frame(refresh_window)
    upper_section.grid(row=0,column=0,sticky='w')
    lower_section = tk.Frame(refresh_window)
    lower_section.grid(row=1,column=0,sticky='w')


    #PERSONALITY

    #Create a frame for Personality, Pandexity and the Submit button.
    pp_frame = ttk.LabelFrame(upper_section)
    pp_frame.grid(column=0,row=0, sticky='nw')

    if current_date == '31 December':
        ny_button = ttk.Button(pp_frame, text="Create New Year")
        ny_button.grid(column=0, row=4, padx=10, pady=10)
        def new_year():
            sheets = wb.sheetnames
            for sheet in sheets:
                sheet1 = wb[sheet]
                sheet2 = wb_year.create_sheet(sheet)
                maxr = sheet1.max_row
                maxc = sheet1.max_column
                for r in range (1, maxr + 1):
                    for c in range (1, maxc + 1):
                        sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
                wb_year.save(f'{current_year}.xlsx')
                os.remove("active.xlsx")
        ny_button.configure(command=new_year)
    #Create a frame to hold the Personality Sliders.
    personality_section = ttk.LabelFrame(pp_frame, text="Personality")
    personality_section.grid(column=0, row=0, padx=10, pady=10)

    #Create some Variables and a function to get the Labels to show the Slider values in Real Time.
    slider_values = [tk.IntVar() for _ in range(6)]
    sliders = [ttk.Scale(personality_section, from_=0, to=10, orient="horizontal", variable=slider_values[i]) for i in range(6)]
    value_labels = [ttk.Label(personality_section, text="0") for i in range(6)]
    def update_labels(event):
        for i, label in enumerate(value_labels):
            label["text"] = "{}".format(slider_values[i].get())
    # Bind the function to the sliders' "movement" events
    for slider in sliders:
        slider.bind("<Motion>", update_labels)


    # Create labels and sliders for each personality parameter
    smart_label = ttk.Label(personality_section, text="Smart").grid(column=0, row=0)
    sliders[0].grid(column=1, row=0)
    value_labels[0].grid(column=2, row=0)

    disciplined_label = ttk.Label(personality_section, text="Disciplined").grid(column=0, row=1)
    sliders[1].grid(column=1, row=1)
    value_labels[1].grid(column=2, row=1)

    meticulous_label = ttk.Label(personality_section, text="Meticulous").grid(column=0, row=2)
    sliders[2].grid(column=1, row=2)
    value_labels[2].grid(column=2, row=2)

    reserved_label = ttk.Label(personality_section, text="Reserved").grid(column=0, row=3)
    sliders[3].grid(column=1, row=3)
    value_labels[3].grid(column=2, row=3)

    calm_label = ttk.Label(personality_section, text="Calm").grid(column=0, row=4)
    sliders[4].grid(column=1, row=4)
    value_labels[4].grid(column=2, row=4)

    kind_label = ttk.Label(personality_section, text="Kind").grid(column=0, row=5)
    sliders[5].grid(column=1, row=5)
    value_labels[5].grid(column=2, row=5)


    # PANDEXITY

    # Create the Pandexity section
    pandexity_section = ttk.LabelFrame(pp_frame, text="Pandexity")
    pandexity_section.grid(column=0, row=2, padx=10, pady=10)

    #Real Time Slider Value
    slider_values_pan = [tk.IntVar() for _ in range(2)]
    sliders_pan = [ttk.Scale(pandexity_section, from_=0, to=10, orient="horizontal", variable=slider_values_pan[i]) for i in range(2)]
    value_labels_pan = [ttk.Label(pandexity_section, text="0") for i in range(2)]
    def update_labels_pan(event):
        for i, label in enumerate(value_labels_pan):
            label["text"] = "{}".format(slider_values_pan[i].get())
    for slider in sliders_pan:
        slider.bind("<Motion>", update_labels_pan)

    # Create labels and sliders for each Pandexity parameter
    happiness_label = ttk.Label(pandexity_section, text="Happiness")
    happiness_label.grid(column=0, row=0)
    sliders_pan[0].grid(column=1, row=0)
    value_labels_pan[0].grid(column=2, row=0)

    pandexity_label = ttk.Label(pandexity_section, text="Pandexity")
    pandexity_label.grid(column=0, row=1)
    sliders_pan[1].grid(column=1, row=1)
    value_labels_pan[1].grid(column=2, row=1)

    # PROJECTS

    project_manager = ttk.LabelFrame(upper_section, text="Project Manager")
    project_manager.grid(column=1,row=0, padx=10, pady=10, sticky='nsew')

    # Create the Projects section
    project_manager_section = ttk.LabelFrame(project_manager)
    project_manager_section.grid(column=0, row=0, padx=10, pady=10, sticky='nsw')
    # Create an empty list to store the activities for this project
    project_activities = []


    # Get the list of Project and Activities in the workbook and create Sections.
    project_frame_full = ttk.LabelFrame(lower_section, text="Projects")
    project_frame_full.grid(column=0,row=0, sticky='nwe')
    projects_list = wb.sheetnames[1:]
    no_of_projects=len(projects_list)
    project_activities = []
    i=-1
    slider_values_activities=[]
    sliders_activities=[]
    value_labels_activities=[]
    for proj in projects_list:
        i=i+1
        j,r=1,-1
        project_frame = tk.LabelFrame(project_frame_full, text=proj)
        project_frame.grid(column=i, row=1, padx=10, pady=10, sticky='n')
        no_of_activities=-1
        for cell in wb[proj].iter_cols(min_row=1, max_row=1):
            no_of_activities += 1

        # create a label for each activity in the project
        slider_values_activities.append([tk.IntVar() for _ in range(no_of_activities)])
        sliders_activities.append([tk.Scale(project_frame, from_=0, to=10, orient="horizontal", variable=slider_values_activities[i][z]) for z in range(no_of_activities)])
        
        for activity in wb[proj].iter_cols(min_row=1, max_row=1, max_col=wb[proj].max_column, min_col=2):
            j=j+1
            r=r+1
            cell = activity[0]
            activity_label = tk.Label(project_frame, text=cell.value)
            activity_label.grid(column=0,row=j)
            sliders_activities[i][r].grid(column=1, row=j)


    # Function to add a new project and its activities
    def add_project():
        # Create a new frame to hold the project and its activities
        project_frame = ttk.Frame(project_manager)
        project_frame.grid(column=0, row=1, padx=10, pady=10, sticky='nw')

        name = project_name_entry_add.get()

        # Create a label for the project name
        project_name_label = ttk.Label(project_frame, text=f"Project {name}")
        project_name_label.grid(column=0, row=0)

        # Create a label for the project activities
        project_activities_label = ttk.Label(project_frame)
        project_activities_label.grid(column=0, row=1)

        # Create an empty list to store the activities for this project
        project_activities = []

        # Function to add a new activity to this project
        def add_activities():

            # Create a label for the activity
            activity_label = ttk.Label(project_frame, text="Activity Name :")
            activity_label.grid(column=0, row=2)

            # Create an entry to enter the activity
            activity_entry = ttk.Entry(project_frame)
            activity_entry.grid(column=2, row=2)

            def add_activity():
                project_activities.append((activity_entry.get()))
                activity_entry.delete(0,300)
            add_activity_button = ttk.Button(project_frame,text="Add", command=add_activity)
            add_activity_button.grid(column=3, row=2)
            
        def save_project():
            proj_sheet = wb.create_sheet(f"Project {name}")
            proj_sheet['A1'] = 'Date'
            if len(project_activities) == 0:
                pass
            else:
                for x in range(len(project_activities)):
                    cell = proj_sheet.cell(row=1, column=x+2)
                    cell.value = project_activities[x]
            wb.save("active.xlsx")       
            project_frame.destroy()
            

            

        # Create a button to add a new activity to this project
        add_activities_button = ttk.Button(project_frame, text="Add Activities", command=add_activities)
        add_activities_button.grid(column=2, row=0)

        save_proj_button = ttk.Button(project_frame, text="Save Project", command=save_project)
        save_proj_button.grid(column=3, row=0)

    selected_proj_var = tk.StringVar(root)
    edit_project_menu = ttk.OptionMenu(project_manager_section, selected_proj_var, *projects_list).grid(column=3,row=1)


    # Function to remove the project and its activities
    def edit_project():
        # Create a label for the project name
        selected_proj = selected_proj_var.get()
        print(selected_proj)
        project_frame = ttk.Frame(project_manager)
        project_frame.grid(column=0, row=1, padx=10, pady=10, sticky='nw')

        project_name_label = ttk.Label(project_frame, text=selected_proj)
        project_name_label.grid(column=0, row=0)

        # Create a label for the project activities
        project_activities_label = ttk.Label(project_frame)
        project_activities_label.grid(column=0, row=1)

        # Create an empty list to store the activities for this project
        project_activities = []
        proj_sheet = wb[selected_proj]

        def del_activities():

            # Create a label for the activity
            activity_label = ttk.Label(project_frame, text="Delete Activity Number:")
            activity_label.grid(column=0, row=2)

            # Create an entry to enter the activity
            activity_entry = ttk.Entry(project_frame)
            activity_entry.grid(column=2, row=2)

            def del_activity():
                col_no = int(activity_entry.get())+1
                proj_sheet.delete_cols(col_no)
                activity_entry.delete(0,300)

            del_activity_button = ttk.Button(project_frame,text="Delete", command=del_activity)
            del_activity_button.grid(column=3, row=2)

        def add_activities():

            # Create a label for the activity
            activity_label = ttk.Label(project_frame, text="Activity Name :")
            activity_label.grid(column=0, row=2)

            # Create an entry to enter the activity
            activity_entry = ttk.Entry(project_frame)
            activity_entry.grid(column=2, row=2)

            def add_activity():
                project_activities.append((activity_entry.get()))
                activity_entry.delete(0,300)
            add_activity_button = ttk.Button(project_frame,text="Add", command=add_activity)
            add_activity_button.grid(column=3, row=2)
        
    

        def save_project():
            if len(project_activities) == 0:
                pass
            else:
                filled_col=-1
                for cell in wb[selected_proj].iter_cols(min_row=1, max_row=1):
                    filled_col += 1
                for x in range(len(project_activities)):
                    cell = proj_sheet.cell(row=1, column=filled_col+x+2)
                    cell.value = project_activities[x]
            wb.save("active.xlsx")       
            project_frame.destroy()
        
        
        def complete_project():
            sheet1 = wb[selected_proj]
            sheet2 = wb_year.create_sheet(selected_proj)
            maxr = sheet1.max_row
            maxc = sheet1.max_column
            for r in range (1, maxr + 1):
                for c in range (1, maxc + 1):
                    sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
            wb_year.save(f'{current_year}.xlsx')
            wb.remove_sheet(sheet1)
            wb.save("active.xlsx")
            project_frame.destroy()

        def delete_project():
            sheet1 = wb[selected_proj]
            wb.remove_sheet(sheet1)
            wb.save("active.xlsx")
            project_frame.destroy()

            

        add_activities_button = ttk.Button(project_frame, text="Add Activities", command=add_activities)
        add_activities_button.grid(column=2, row=0)

        del_activities_button = ttk.Button(project_frame, text="Delete Activities", command=del_activities)
        del_activities_button.grid(column=3, row=0)

        save_proj_button = ttk.Button(project_frame, text="Save Project", command=save_project)
        save_proj_button.grid(column=6, row=0)

        comp_proj_button = ttk.Button(project_frame, text="Complete Project", command=complete_project)
        comp_proj_button.grid(column=4, row=0)

        del_proj_button = ttk.Button(project_frame, text="Delete Project", command=delete_project)
        del_proj_button.grid(column=5, row=0)



    # Create an entry to enter the project name
    project_name_entry_add = ttk.Entry(project_manager_section)
    project_name_entry_add.grid(column=0, row=1)

    # Create a button to add a new project
    add_project_button = ttk.Button(project_manager_section, text="Add Project", command=add_project)
    add_project_button.grid(column=1, row=1)

    empt = tk.Label(project_manager_section, text='         ')
    empt.grid(column=2,row=1)

    
    empt = tk.Label(project_manager_section, text='         ')
    empt.grid(column=5,row=1)
    # Create a Tkinter variable to store the selected sheet name


        
    # Create a button to remove this project and its activities
    edit_project_button = ttk.Button(project_manager_section, text="Edit Project", command=edit_project)
    edit_project_button.grid(column=4, row=1)



                    
    def save_scores():
        # Define the data to be added to the sheet
        current_date = datetime.now()
        date = current_date.date()
        smart = round(sliders[0].get())
        disciplined = round(sliders[1].get())
        meticulous = round(sliders[2].get())
        reserved = round(sliders[3].get())
        calm = round(sliders[4].get())
        kind = round(sliders[5].get())
        average = round((smart+disciplined+meticulous+reserved+calm+kind)/6)
        pandexity = round(sliders_pan[1].get())
        happiness = round(sliders_pan[0].get())
        # Add the data to the sheet
        per_sheet.append([date, smart, disciplined, meticulous, reserved, calm, kind, average, pandexity, happiness])
        

        i=-1
        for sheet in projects_list:
            active_sheet = wb[sheet]
            i=i+1
            no_of_activities=-1
            for cell in wb[sheet].iter_cols(min_row=1, max_row=1):
                no_of_activities += 1
        
            active_sheet.append([date] + [sliders_activities[i][r].get() for r in range(no_of_activities)])
        wb.save("active.xlsx")

    # Create a button to save all scores to the Excel file
    save_button = ttk.Button(pp_frame, text="Submit", command=save_scores)
    save_button.grid(column=0, row=3, padx=10, pady=10)

    #TOOLS

    graphs_list = ['bar chart', 'pie chart', 'line chart']

    path = r'*.xlsx'
    year_list = glob.glob(path)
    

    graphs_frame=ttk.LabelFrame(upper_section, text="Graphs")
    graphs_frame.grid(row=0,column=2, padx=10, pady=10, sticky='nse')

    graph_selected_year_var = tk.StringVar(root)
    graph_year_menu = ttk.OptionMenu(graphs_frame, graph_selected_year_var, *year_list)
    graph_year_menu.grid(column=0,row=1)
    
    def graph_select_project():
        graph_selected_year = graph_selected_year_var.get()
        temp_frame = tk.Frame(graphs_frame)
        temp_frame.grid(column=0, row=3, columnspan=2)

        if graph_selected_year == "active.xlsx":
            graph_projects_list = wb.sheetnames
    
        else:
            owb = xl.load_workbook(graph_selected_year)
            graph_projects_list= owb.sheetnames

        graph_selected_proj_var = tk.StringVar(root)
        graph_project_menu = ttk.OptionMenu(temp_frame, graph_selected_proj_var, *graph_projects_list)
        graph_project_menu.grid(column=0,row=0, sticky='w') 

        def graph_select_activity():
            graph_selected_proj = graph_selected_proj_var.get()
            df = pd.read_excel(r'{}'.format(graph_selected_year), sheet_name=f'{graph_selected_proj}')
            act_list = list(df.columns)
            act_list = act_list[1:]
            
            graph_selected_activity_var = tk.StringVar(root)
            graph_activity_menu = ttk.OptionMenu(temp_frame, graph_selected_activity_var, *act_list)
            graph_activity_menu.grid(row=1,column=0)

            def create_graph_final():
                graph_selected_activity = graph_selected_activity_var.get()
                plt.plot(df["Date"],df[graph_selected_activity])
                plt.show()
                temp_frame.destroy()
            final_button = ttk.Button(temp_frame, text="Create", command=create_graph_final)
            final_button.grid(column=1,row=1, sticky='e')

        chart_button2 = ttk.Button(temp_frame, text="Select Project", command=graph_select_activity)
        chart_button2.grid(column=1,row=0, stick='e')

    chart_button1 = ttk.Button(graphs_frame, text="Select Year", command=graph_select_project)
    chart_button1.grid(column=1,row=1, sticky='e')

    #ANALYSIS SECTION {MACHINE LEARNING}


    analysis_frame=ttk.LabelFrame(upper_section, text="Analysis")
    analysis_frame.grid(row=0,column=3, padx=10, pady=10, sticky='nse')

    def analyze():
        dependent_variable = "Happiness"
        independent_variables = ['Smart', 'Disciplined', 'Meticulous', 'Reserved', 'Calm', 'Kind', 'Pandexity']

        df = pd.read_excel(r'active.xlsx', sheet_name='Personality')

        # Create a linear regression model using the `statsmodels` library
        model = sm.OLS(df[dependent_variable], df[independent_variables])

        # Fit the model to the data
        results = model.fit()

        # Print the summary of the model
        print(results.summary())

        global pop
        pop = tk.Toplevel(root)
        pop.title("Analysis Summary")
        pop.geometry("640x480")
        # Create a Label Text
        label = tk.Label(pop, text=results.summary())
        label.grid(row=0, column=0)

    analyze_button = ttk.Button(analysis_frame, text="Analyze", command=analyze)
    analyze_button.grid(column=0,row=0)
    
    def refresh():
        refresh_window.destroy()
        run()

    refresh_button = ttk.Button(project_manager_section, text='Refresh', command=refresh)
    refresh_button.grid(row=1,column=6)


run()

# Run the main loop
root.mainloop()

